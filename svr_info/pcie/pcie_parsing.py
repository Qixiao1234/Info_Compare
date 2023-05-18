import subprocess
import os
import sys
import time
import re

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill

from openpyxl.styles.borders import Border, Side

filename = "dump.txt"

no_value = "no value"

def dump_info(bus_id):

    cmd = 'lspci -vvs '+bus_id

    output = subprocess.check_output(cmd, shell=True)
    try:
        output = output.decode()
    except UnicodeDecodeError as err:
        print("[Decode Error] ", err)
        print("Replace byte '\\x89' from <%s>" % bus_id)
        output = output.replace(b'\x89',b'')
        output = output.decode()

    if output == "":
        return False

    with open(filename, 'a+') as file:
        file.write('> '+cmd+':\n')
        file.write(output+'\n')
    file.close()

    output = output.split('\n')

    for line in output:
        if bus_id in line:
            line = line.split(bus_id)[1]
            pcie_name = line
        if 'LnkSta' in line:

            match = re.search(r"Speed\s*(.+?)GT\/s,\s*Width\s*x(\d+)", line)
            if match:
                line = match.group(0)
                pcie_status = line
    

    try: pcie_status
    except NameError: pcie_status = no_value

    return pcie_name, pcie_status    

def find_bus_id(input_line):

    parsing = input_line.split('-[')
    #return ['+', '0000:3a]-+-00.0', '3b-42]----00.0', '3c-42]--+-00.0', '3d-3e]----00.0']

    parsing_list = [item for item in parsing if re.search(r"(..?)-(..?)]", item)]
    #return ['3b-42]----00.0', '3c-42]--+-00.0', '3d-3e]----00.0']


    if not len(parsing_list) == 0:
        for item in parsing_list:
            has_sub_tree = False

            match = re.search(r"(\w\w?)-", item) #item = "3b-42]----00.0"
            if match:
                bus_id_prefix = match.group(1) #3b

            match = item.split(']')[1] # ----00.0
            if '+' in match:
                has_sub_tree = True
            for rule in (("-", ""), ("+", "")):
                bus_id_suffix = match.replace(*r) # 00.0
            bus_id = bus_id_prefix+':'+ bus_id_suffix #3b:00.0

def find_bus_info(bus_id, info):

    cmd = 'lspci -vs '+bus_id

    output = subprocess.getoutput(cmd)

    if info == 'check':
        if output == "":
            return False
        else:
            return True

    with open(filename, 'a+') as file:
        file.write('> '+cmd+':\n')
        file.write(output+'\n')
    file.close()

    output = output.split('\n')

    for line in output:
        if 'Bus:' in line:
            if info == 'primary':
                match = re.search(r"primary=(..?)", line)
                if match:
                    match = match.group(1)
            elif info == 'secondary':
                match = re.search(r"secondary=(..?)", line)
                if match:
                    match = match.group(1)
        if 'Flags' in line:
            if info == 'NUMA':
                match = re.search(r"node\s(.?)", line)
                if match:
                    match = str(match.group(1))
                else:
                    #for single socket, no NUMA option
                    match = '0'

    try:
        return match
    except UnboundLocalError:
        return False
 


def bus_id_check(bus_id):

    cmd = 'lspci -vs '+bus_id

    output = subprocess.getoutput(cmd)

    if output == "":
        return False

def fill_info_to_cell(ws, col_count, row_count, bus_id):

    if type(bus_id) == list:
        bus_id = bus_id[0]

    if False == dump_info(bus_id):
        bus_input = bus_id+'\n'+no_value
    else:
        pcie_name, pcie_status = dump_info(bus_id)
        bus_input = bus_id+'\n'+pcie_name+'\n'+pcie_status

     

    while True:

        target_cell =ws[chr(65+col_count)+str(row_count)].value
        if not 'None' in str(target_cell):
            row_count = row_count + 1
            continue
        else:
            ws[chr(65+col_count)+str(row_count)] = bus_input
            break


def get_curr_coordinate(ws, bus_id):

    if type(bus_id) == list:
        bus_id = bus_id[0]

    node_exist = False

    for row in ws:
        for cell in row:
            value = str(cell.value)
            if not 'None' in value and not 'N/A' in value:
                if bus_id in value:
                    cell_col = cell.coordinate[0]
                    cell_col = ord(cell_col)
                    cell_row = cell.row
                    node_exist = True
                    return cell_col-65, cell_row
    if node_exist == False:
        return False

def fill_child_value(ws, col_count, row_count, bus_id, socket):

    sub_id = pcie_node[socket][bus_id]

def adjust_datasheet(wb, ws):

    #-----Because there are so many empty rows between root PCIE nodes, firstly copy those none empty values to another sheet, called 'PCIE_back' 

    row_boundary = []
    row_remove = []

    for row in ws:
        for cell in row:
            value = str(cell.value)
            cell_row = re.findall(r'\d+', cell.coordinate)[0]
            if 'None' in value or 'N/A' in value:
                row_empty = True
            else:
                row_empty = False
                count = 0
                break
        
        if row_empty == True:
            if cell_row not in row_boundary and count == 0:
                row_boundary.append(cell_row)
            if cell_row not in row_remove and count > 0:
                row_remove.append(cell_row)
            count = count + 1

    
    wb.create_sheet('PCIE_back')

    ws2 = wb['PCIE_back']

    ws2_row_boundary = []
    ws2_row_count = 1


    for row in ws:
        for cell in row:
            cell_col = cell.coordinate[0]
            cell_row = re.findall(r'\d+', cell.coordinate)[0]
           
            if not cell_row in row_remove:
                ws2[cell_col+str(ws2_row_count)].value = cell.value
                row_check = True

                if cell_row in row_boundary:
                    ws2_row_boundary.append(ws2_row_count)
            else:
                row_check = False
                break
        if row_check == True:
            ws2_row_count = ws2_row_count + 1

    #-------After migrate the values, remove original sheet, and rename 'PCIE_back' to original 'PCIE', then start the styling
    wb.remove(wb['PCIE'])
    ws = wb['PCIE_back']
    ws.title = 'PCIE'

    ##----Styling----##
    max_colum = 0

    for row in ws:
        for cell in row:
            value = str(cell.value)
            cell_col = cell.coordinate[0]
            cell_col = ord(cell_col)
            if not 'None' in value and not 'N/A' in value:
                if cell_col > max_colum:
                    max_colum = cell_col

                cell.alignment = Alignment(wrapText=True)

    for i in range(66, max_colum+1):
        ws.column_dimensions[chr(i)].width = 50

        
    ###----painting the boundary between root PCIE nodes
    for row_num in ws2_row_boundary:
        for i in range(65, max_colum+1):
            ws[chr(i)+str(row_num)].fill = PatternFill(start_color='002060', end_color='002060', fill_type = "solid")



def input_excel(Reports_path, root_pcie_bridge_list, pcie_node, pcie_layer):

    if Reports_path == "None":

        wb = Workbook(write_only=True)
        Reports_path = "Result.xlsx"
        wb.save(Reports_path)

    wb = openpyxl.load_workbook(Reports_path)

    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    elif 'PCIE' in wb.sheetnames:
        wb.remove(wb['PCIE'])

    wb.create_sheet('PCIE')

    ws = wb['PCIE']

    row_interval = 100
    row_count = 0
    row_botton = 0


    for curr_socket in pcie_layer:
        pcie_node_list = list(pcie_node[curr_socket].values())

        row_count = row_botton + 1
        col_count = 0
        ws[chr(65+col_count)+str(row_count)] = curr_socket

        row_count = row_count + 1

        for root_bus_id in pcie_layer[curr_socket][1]:
            col_count = 1
            fill_info_to_cell(ws, col_count, row_count, root_bus_id)
            row_count = row_count + row_interval
            row_botton = row_count

        for layer in pcie_layer[curr_socket]:
            if not layer == 1:
                for bus_id in pcie_layer[curr_socket][layer]:

                    parent_bus_id = [k for k,v in pcie_node[curr_socket].items() if bus_id in v]
                    col_count, row_count = get_curr_coordinate(ws, parent_bus_id)

                    col_count = col_count + 1
                    fill_info_to_cell(ws, col_count, row_count, bus_id)


    adjust_datasheet(wb, ws)

    #before closing file, make the default page to Readme
    wb.active = 0    

    wb.save(Reports_path)

def main_binary(path_info):

    Reports_path = path_info

    main_processing(Reports_path)

def main():

    Reports_path = "None"

    main_processing(Reports_path)

def main_processing(Reports_path):

    host = 'dut'
    cmdict = {}
    cmdict[host] = {}

    #store cmd from filename
    with open(filename, "r") as file:
        for line in file:
            if line != "\n":
                if line.startswith("> "):
                    cmd = line[2:].split(":")[0]
                    cmdict[host][cmd] = []

                else:
                    cmdict[host][cmd].append(line)

    pcie_node = {}
    single_socket = False
    #for single socket
    if len(cmdict[host]["dmesg | grep 'NUMA node'"]) == 0:
        pcie_node['Socket 0'] = {}
        single_socket = True
    else:
        #dual sockets
        pcie_node['Socket 0'] = {}
        pcie_node['Socket 1'] = {}
        pcie_node['0'] = []
        pcie_node['1'] = []

        for line in cmdict[host]["dmesg | grep 'NUMA node'"]:
            if 'node 0' in line:
                bus_id = line.split(':')[1]
                pcie_node['0'].append(bus_id)
            elif 'node 1' in line:
                bus_id = line.split(':')[1]
                pcie_node['1'].append(bus_id)

    for line in cmdict[host]["lspci | grep Root"]:
        bus_id = line.split(' ')[0] #17:00.0
        bus_secondary = find_bus_info(bus_id, 'secondary') + ':00.0'
        if single_socket:
            pcie_node['0'][bus_id] = bus_secondary
        else:
            bus_id_short = bus_id.split(':')[0] #17
            for item in pcie_node['0']:
                if item == bus_id_short:
                    pcie_node['Socket 0'][bus_id] = bus_secondary
            for item in pcie_node['1']:
                if item == bus_id_short:
                    pcie_node['Socket 1'][bus_id] = bus_secondary

    #{'Socket 0': {'00:1c.0': {}, '17:00.0': {}, '3a:00.0': {}, '5d:02.0': {}, '5d:03.0': {}}, 'Socket 1': {'ae:00.0': {}, 'ae:02.0': {}, 'd7:00.0': {}}}

    parsing_list = cmdict[host]["lspci | grep 'PCI bridge' | grep -v 'Root'"]

    no_root_pcie_bridge_list = [item.split()[0] for item in parsing_list] #['01:00.0', '3b:00.0', '3c:00.0', '3c:01.0', '3c:03.0', 'd8:00.0', 'd9:02.0']

    for bus_id in no_root_pcie_bridge_list:
        socket_num = find_bus_info(bus_id, 'NUMA')
        pcie_node['Socket '+socket_num][bus_id] = {}

        bus_secondary = find_bus_info(bus_id, 'secondary')

        match_list = []
        for bus_id_search in no_root_pcie_bridge_list:
            bus_id_search_match = bus_id_search.split(':')[0]
            if bus_id_search_match == bus_secondary:
                match_list.append(bus_id_search)
        if not len(match_list) == 0:
            pcie_node['Socket '+socket_num][bus_id] = match_list
        else:
            pcie_node['Socket '+socket_num][bus_id] = bus_secondary+':00.0'

    #for trying dump end device info

    if 'Socket 1' in pcie_node:


        list_key_node0 = list(pcie_node['Socket 0'].keys())
        list_key_node1 = list(pcie_node['Socket 1'].keys())

        list_value_node0 = []
        list_value_node1 = []
        for value in pcie_node['Socket 0'].values():
            if not type(value) == list:
                list_value_node0.append(value)
            else:
                for item in value:
                    list_value_node0.append(item)

        for value in pcie_node['Socket 1'].values():
            if not type(value) == list:
                list_value_node1.append(value)
            else:
                for item in value:
                    list_value_node1.append(item)

        list_node0 = [item for item in list_value_node0 if item not in list_key_node0]
        list_node1 = [item for item in list_value_node1 if item not in list_key_node1]

    else:

        list_key_node0 = list(pcie_node['Socket 0'].keys())

        list_value_node0 = []
        for value in pcie_node['Socket 0'].values():
            if not type(value) == list:
                list_value_node0.append(value)
            else:
                for item in value:
                    list_value_node0.append(item)

        list_node0 = [item for item in list_value_node0 if item not in list_key_node0]
    
    for bus_id in list_node0: #['18:00.0', '5e:00.0', '5f:00.0', '02:00.0', '3d:00.0', '3f:00.0', '41:00.0']
        bus_id_ori = bus_id
        bus_id_prefix = bus_id.split('.')[0] # 18:00
        bus_id_count = int(bus_id.split('.')[1]) # 0


        end_device_id_list = []
        end_device_id_list.append(bus_id_ori)
        while True:
            bus_id_count = bus_id_count+1
            bus_id = bus_id_prefix+'.'+str(bus_id_count)
            if False == bus_id_check(bus_id):
                break
            else:
                end_device_id_list.append(bus_id)

        if len(end_device_id_list) > 1: #['41:00.0', '41:00.1', '41:00.2', '41:00.3']

            for key, value in pcie_node['Socket 0'].items(): 
                #to replace value
                if bus_id_ori == value: 
                    pcie_node['Socket 0'][key] = end_device_id_list

    if 'Socket 1' in pcie_node:
        for bus_id in list_node1: #['18:00.0', '5e:00.0', '5f:00.0', '02:00.0', '3d:00.0', '3f:00.0', '41:00.0']
            bus_id_ori = bus_id
            bus_id_prefix = bus_id.split('.')[0] # 18:00
            bus_id_count = int(bus_id.split('.')[1]) # 0


            end_device_id_list = []
            end_device_id_list.append(bus_id_ori)
            while True:
                bus_id_count = bus_id_count+1
                bus_id = bus_id_prefix+'.'+str(bus_id_count)
                if False == bus_id_check(bus_id):
                    break
                else:
                    end_device_id_list.append(bus_id)

            if len(end_device_id_list) > 1: #['41:00.0', '41:00.1', '41:00.2', '41:00.3']

                for key, value in pcie_node['Socket 1'].items(): 
                    #to replace value
                    if bus_id_ori == value: 
                        pcie_node['Socket 1'][key] = end_device_id_list

    parsing_list = cmdict[host]["lspci | grep Root"]
    if len(parsing_list) == 0:
        parsing_list = cmdict[host]["lspci | grep 'PCI bridge' | grep -v 'Root'"]
    root_pcie_bridge_list = [item.split()[0] for item in parsing_list]


    ##try to classify the bus id to different layer

    pcie_layer = {}

    for bus_id in root_pcie_bridge_list:
        count = 1

        socket_num = find_bus_info(bus_id, 'NUMA')
        if not 'Socket '+socket_num in pcie_layer:
            pcie_layer['Socket '+socket_num] = {}
        if not count in pcie_layer['Socket '+socket_num]:
            pcie_layer['Socket '+socket_num][count] = []

        pcie_layer['Socket '+socket_num][count].append(bus_id)
        while True:
            if False == find_bus_info(bus_id, 'secondary'):
                break
            count = count + 1
            if not count in pcie_layer['Socket '+socket_num]:
                pcie_layer['Socket '+socket_num][count] = []
            
            bus_secondary = find_bus_info(bus_id, 'secondary')
            bus_id = bus_secondary + ':00.0'
            if True == find_bus_info(bus_id, 'check'):
                pcie_layer['Socket '+socket_num][count].append(bus_id)

            for bus_id_search in no_root_pcie_bridge_list:
                bus_id_search_prefix = bus_id_search.split(':')[0]
                if bus_id_search_prefix == bus_secondary:
                    if not bus_id == bus_id_search: #to prevent duplicate
                        pcie_layer['Socket '+socket_num][count].append(bus_id_search)


    for socket in pcie_layer.copy(): #use copy(), because the dict might change during iteration 
        for layer in pcie_layer[socket].copy():
            for bus_id in pcie_layer[socket][layer].copy():
                if bus_id in pcie_node[socket]: 
                    bus_id = pcie_node[socket][bus_id]

                    #-+-[0000:d7]-+-00.0-[d8-db]----00.0-[d9-db]----02.0-[da-db]----00.0
                    #in this case, because 'd9:02.0', the layer 4 will not be auto generated thru
                    #fomer process, which only try to search like 'd9:00.0' child node. 
                    if not layer+1 in pcie_layer[socket]:
                        pcie_layer[socket][layer+1] = [] 

                    if not bus_id in pcie_layer[socket][layer+1].copy():
                        if type(bus_id) == list:
                            for item in bus_id:
                                if not item in pcie_layer[socket][layer+1].copy():
                                    pcie_layer[socket][layer+1].append(item)
                        else:
                            pcie_layer[socket][layer+1].append(bus_id)

    input_excel(Reports_path, root_pcie_bridge_list, pcie_node, pcie_layer)

def parsing_arg():
    if len(sys.argv()) > 1:
        path_info = sys.argv[1]
    else:
        path_info = "None"

    return path_info

if __name__ == "__main__":
    main()