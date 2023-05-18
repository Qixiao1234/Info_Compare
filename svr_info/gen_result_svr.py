#! /usr/bin/python
import re, os, sys, argparse
from string import Template
from collections import defaultdict
from collections import OrderedDict

import openpyxl
import re

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from openpyxl.styles import Font
from openpyxl.styles import colors
from openpyxl.chart import BarChart, Series, Reference
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.styles import Alignment
import time
import pipes

from svrinfo_extend import Svrinfo
import yaml

with open('loadnum.txt', 'r+', encoding='utf-8') as a:
    run_num = a.readlines()[-1]
    # print(int(run_num.strip()))
    # a.write(str(int(run_num) + 1) + '\n')
work_dir = os.getcwd()
Reports_path = str(int(run_num.strip())) + "_Reports.xlsx"


# print(Reports_path)

##############Report out###############################

def adding_sheet(sheet_name):
    if os.path.exists(Reports_path) == False:
        gen_excel()

    # load excel
    global xfile
    xfile = openpyxl.load_workbook(Reports_path)
    if sheet_name in xfile.sheetnames:
        xfile.remove(xfile[sheet_name])

    if sheet_name == "Legal":
        xfile.create_sheet(sheet_name, 0)
    elif sheet_name == "ServerInfo":
        xfile.create_sheet(sheet_name, 1)
    else:
        xfile.create_sheet(sheet_name)

    if 'Sheet' in xfile.sheetnames:
        xfile.remove(xfile['Sheet'])

    # save file
    xfile.save(Reports_path)


def gen_excel():
    # generate excel file
    wb = Workbook(write_only=True)
    wb.save(Reports_path)


def gen_report_net(title, table_name, dut, golden):
    list_item = []
    list_golden = []
    for k, v in golden.items():
        for k, v in v.items():
            list_item.append(k)
            list_golden.append(v)

    list_dut = []

    for k, v in dut.items():
        for k, v in v.items():
            list_dut.append(v)

    # In order to record the scope of table
    global table_start
    global table_end
    global last_row

    Get_input_row()

    worksheet['A' + str(last_row)].font = Font(size=14, bold=True)
    worksheet['A' + str(last_row)] = title
    last_row = last_row + 1
    table_start = last_row
    worksheet['A' + str(last_row)] = table_name
    worksheet['B' + str(last_row)] = 'dut'
    worksheet['C' + str(last_row)] = 'golden'

    last_row = last_row + 1

    row_input = last_row

    for i in list_item:
        worksheet['A' + str(row_input)] = i
        row_input = row_input + 1

    row_input = last_row
    for i in list_dut:

        all_item = ""
        for count in range(len(i)):
            item = i[count]
            if count == 0:
                all_item = item
            else:
                all_item = all_item + ', ' + item

        worksheet['B' + str(row_input)].alignment = Alignment(wrapText=True)

        worksheet['B' + str(row_input)] = all_item
        row_input = row_input + 1

    row_input = last_row
    for i in list_golden:
        all_item = ""
        for count in range(len(i)):
            item = i[count]
            if count == 0:
                all_item = item
            else:
                all_item = all_item + ', ' + item

        worksheet['C' + str(row_input)].alignment = Alignment(wrapText=True)
        worksheet['C' + str(row_input)] = all_item
        row_input = row_input + 1

    table_end = row_input - 1

    adjust()


def gen_report_graph(title, dut, golden):
    global last_row, table_start, table_end
    Get_input_row()

    worksheet['A' + str(last_row)] = title
    last_row = last_row + 1
    table_start = last_row

    worksheet['A' + str(last_row)].font = Font(color=colors.WHITE)
    worksheet['B' + str(last_row)].font = Font(color=colors.WHITE)
    worksheet['C' + str(last_row)].font = Font(color=colors.WHITE)
    worksheet['D' + str(last_row)].font = Font(color=colors.WHITE)
    worksheet['A' + str(last_row)] = 'dut'
    worksheet['B' + str(last_row)] = 'dut'
    worksheet['C' + str(last_row)] = 'golden'
    worksheet['D' + str(last_row)] = 'golden'

    last_row = last_row + 1

    row_input = last_row

    for k, v in dut.items():
        v = sorted(v, reverse=True)
        for i in v:
            worksheet['A' + str(row_input)].font = Font(color=colors.WHITE)
            worksheet['B' + str(row_input)].font = Font(color=colors.WHITE)
            worksheet['A' + str(row_input)] = int(i[2])
            worksheet['B' + str(row_input)] = int(i[1])
            row_input = row_input + 1

    row_input = last_row
    for k, v in golden.items():
        v = sorted(v, reverse=True)
        for i in v:
            worksheet['C' + str(row_input)].font = Font(color=colors.WHITE)
            worksheet['D' + str(row_input)].font = Font(color=colors.WHITE)
            worksheet['C' + str(row_input)] = int(i[2])
            worksheet['D' + str(row_input)] = int(i[1])
            row_input = row_input + 1

    table_end = row_input - 1

    draw_graph()


def draw_graph():
    # setup the chart
    chart = ScatterChart()
    chart.height = 10.5  # default is 7.5
    chart.width = 27.6  # default is 15

    chart.x_axis.title = 'Bandwidth (GB/s)'
    chart.y_axis.title = 'Latency (ns)'

    dut_x_value = Reference(worksheet, min_col=1, min_row=last_row, max_row=table_end)
    dut_y_value = Reference(worksheet, min_col=2, min_row=last_row, max_row=table_end)
    golden_x_value = Reference(worksheet, min_col=3, min_row=last_row, max_row=table_end)
    golden_y_value = Reference(worksheet, min_col=4, min_row=last_row, max_row=table_end)

    series = Series(dut_y_value, dut_x_value, title="dut")
    chart.append(series)

    series = Series(golden_y_value, golden_x_value, title="golden")
    chart.append(series)
    chart.type = "col"
    chart.style = 10
    chart.shape = 4

    worksheet.add_chart(chart, 'A' + str(last_row - 1))


def parse_dimm_info(device):
    global socket_list, channel_list, slot_list
    global dimm_list1, dimm_list2

    socket_list = []
    channel_list = []
    slot_list = []
    dimm_list1 = []
    dimm_list2 = []

    for k, v in device.items():
        for i in v:

            val = i.setdefault('Socket')
            socket_list.append(val)
            val = i.setdefault('Channel')
            channel_list.append(val)
            val = i.setdefault('Slot')
            slot_list.append(val)
            a = i.setdefault('Size')
            b = i.setdefault('Speed')
            c = i.setdefault('Type')
            d = i.setdefault('Rank')
            e = i.setdefault('Manufacturer')
            f = i.setdefault('Part')
            if d == '1':
                d = 'single-rank'
            elif d == '2':
                d = 'dual-rank'
            elif d == '4':
                d = 'quad-rank'
            dimm_list1.append(a + '@' + b + ' ' + c + ' ' + d)
            dimm_list2.append(e + ' ' + f)

    socket_list = sorted(set(socket_list))
    channel_list = sorted(set(channel_list))
    slot_list = sorted(set(slot_list))


def gen_report_dimm(title, dut, golden):
    global last_row, table_start, table_end
    Get_input_row()

    worksheet['A' + str(last_row)] = title
    last_row = last_row + 1
    table_start = last_row
    worksheet['A' + str(last_row)] = 'dut'
    worksheet['B' + str(last_row)] = 'golden'

    last_row = last_row + 1

    row_input = last_row

    parse_dimm_info(dut)
    l = 0

    for i in socket_list:
        worksheet['A' + str(row_input)] = 'Socket ' + str(i)
        row_input = row_input + 1
        for j in channel_list:
            for k in slot_list:
                if "No Module" in dimm_list1[l]:
                    worksheet['A' + str(row_input)] = 'Channel ' + str(j) + ' Slot ' + str(k) + '- No DIMM'
                    row_input = row_input + 1
                    l = l + 1
                else:
                    worksheet['A' + str(row_input)].alignment = Alignment(wrapText=True)
                    worksheet['A' + str(row_input)] = 'Channel ' + str(j) + ' Slot ' + str(k) + ' - ' + dimm_list1[
                        l] + '\n' + dimm_list2[l]
                    row_input = row_input + 1
                    l = l + 1
    dut_row_total = row_input

    row_input = last_row
    parse_dimm_info(golden)
    l = 0

    for i in socket_list:
        worksheet['B' + str(row_input)] = 'Socket ' + str(i)
        row_input = row_input + 1
        for j in channel_list:
            for k in slot_list:
                if "No Module" in dimm_list1[l]:
                    worksheet['B' + str(row_input)] = 'Channel ' + str(j) + ' Slot ' + str(k) + '- No DIMM'
                    row_input = row_input + 1
                    l = l + 1
                else:
                    worksheet['B' + str(row_input)].alignment = Alignment(wrapText=True)
                    worksheet['B' + str(row_input)] = 'Channel ' + str(j) + ' Slot ' + str(k) + ' - ' + dimm_list1[
                        l] + '\n' + dimm_list2[l]
                    row_input = row_input + 1
                    l = l + 1

    golden_row_total = row_input

    if dut_row_total > golden_row_total:
        table_end = dut_row_total - 1
    else:
        table_end = golden_row_total - 1

    adjust_dimm()


def gen_report(title, table_name, dut, golden):
    list_item = []
    list_golden = []

    for k, v in golden.items():
        for k, v in v.items():
            list_item.append(k)
            list_golden.append(v)

    list_dut = []

    for k, v in dut.items():
        for k, v in v.items():
            list_dut.append(v)

    # In order to record the scope of table
    global table_start
    global table_end
    global last_row

    Get_input_row()

    worksheet['A' + str(last_row)].font = Font(size=14, bold=True)
    worksheet['A' + str(last_row)] = title
    last_row = last_row + 1
    table_start = last_row
    worksheet['A' + str(last_row)] = table_name
    worksheet['B' + str(last_row)] = 'dut'
    worksheet['C' + str(last_row)] = 'golden'

    last_row = last_row + 1

    row_input = last_row

    for i in list_item:
        worksheet['A' + str(row_input)] = i
        row_input = row_input + 1

    row_input = last_row
    for i in list_dut:
        worksheet['B' + str(row_input)] = i
        row_input = row_input + 1

    row_input = last_row
    for i in list_golden:
        worksheet['C' + str(row_input)] = i
        row_input = row_input + 1

    table_end = row_input - 1

    adjust()


def gen_report_single_result(title, table_name, dut, golden):
    list_item = []
    list_golden = []

    for k, v in golden.items():
        list_item.append(k)
        v = ''.join(str(x).replace('\n', '') for x in v)
        list_golden.append(v)

    list_dut = []

    for k, v in dut.items():
        v = ''.join(str(x).replace('\n', '') for x in v)
        list_dut.append(v)

    # In order to record the scope of table
    global table_start
    global table_end
    global last_row

    Get_input_row()

    worksheet['A' + str(last_row)].font = Font(size=14, bold=True)
    worksheet['A' + str(last_row)] = title
    last_row = last_row + 1
    table_start = last_row
    worksheet['A' + str(last_row)] = table_name
    worksheet['B' + str(last_row)] = 'dut'
    worksheet['C' + str(last_row)] = 'golden'

    last_row = last_row + 1

    row_input = last_row
    for i in list_dut:
        worksheet['B' + str(row_input)] = i
        worksheet['B' + str(row_input)].alignment = Alignment(wrapText=True)
        row_input = row_input + 1

    row_input = last_row
    for i in list_golden:
        worksheet['C' + str(row_input)] = i
        worksheet['C' + str(row_input)].alignment = Alignment(wrapText=True)
        row_input = row_input + 1

    table_end = row_input - 1

    adjust(mark_different=False)


def adjust(mark_different=True):
    col_width = []

    i = 0

    for col in worksheet.columns:
        for j in range(len(col)):
            if j == 0:
                col_width.append(len(str(col[j].value)))
            else:
                if col_width[i] < len(str(col[j].value)):
                    col_width[i] = len(str(col[j].value))
        i = i + 1

    for i in range(len(col_width)):
        col_letter = get_column_letter(i + 1)
        if col_width[i] > 100:
            worksheet.column_dimensions[col_letter].width = 80
        elif col_width[i] > 10:
            worksheet.column_dimensions[col_letter].width = col_width[i] + 2

    global table_count
    tab = Table(displayName="Table" + str(table_count), ref="A" + str(table_start) + ":C" + str(table_end))
    table_count = table_count + 1
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium15", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    worksheet.add_table(tab)

    # Mark the different parts
    if mark_different == True:
        for row in worksheet[table_start + 1:table_end]:
            for cell in row:
                if worksheet['B' + str(cell.row)].value != worksheet['C' + str(cell.row)].value:
                    worksheet['B' + str(cell.row)].font = Font(color=colors.RED, bold=True)
                    worksheet['C' + str(cell.row)].font = Font(color=colors.RED, bold=True)


def adjust_dimm():
    col_width = []

    i = 0

    for col in worksheet.columns:
        for j in range(len(col)):
            if j == 0:
                col_width.append(len(str(col[j].value)))
            else:
                if col_width[i] < len(str(col[j].value)):
                    col_width[i] = len(str(col[j].value))
        i = i + 1

    for i in range(len(col_width)):
        col_letter = get_column_letter(i + 1)
        if col_width[i] > 100:
            worksheet.column_dimensions[col_letter].width = 80
        elif col_width[i] > 10:
            worksheet.column_dimensions[col_letter].width = col_width[i] + 2

    global table_count
    tab = Table(displayName="Table" + str(table_count), ref="A" + str(table_start) + ":B" + str(table_end))
    table_count = table_count + 1
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium15", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    worksheet.add_table(tab)

    # Mark the different parts
    for row in worksheet[table_start + 1:table_end]:
        for cell in row:
            if worksheet['A' + str(cell.row)].value != worksheet['B' + str(cell.row)].value:
                worksheet['A' + str(cell.row)].font = Font(color=colors.RED, bold=True)
                worksheet['B' + str(cell.row)].font = Font(color=colors.RED, bold=True)
            if worksheet['A' + str(cell.row)].value != None:
                if worksheet['A' + str(cell.row)].value.startswith('Socket'):
                    worksheet['A' + str(cell.row)].font = Font(color=colors.BLACK, bold=True)
            if worksheet['B' + str(cell.row)].value != None:
                if worksheet['B' + str(cell.row)].value.startswith('Socket'):
                    worksheet['B' + str(cell.row)].font = Font(color=colors.BLACK, bold=True)


def Get_input_row():
    global last_row
    # Get input row index
    if worksheet['A1'].value is None:
        last_row = 1
    else:
        for row in worksheet[worksheet.min_row:worksheet.max_row]:
            for cell in row:
                if cell.value is not None:
                    last_row = cell.row
        last_row = last_row + 2


def gen_legal_text():
    xfile = openpyxl.load_workbook(Reports_path)
    worksheet = xfile['Legal']

    text = "Disclosed under and subject to the terms of the CNDA in effect between you and Intel.This document is provided to you solely for your reference, without warranties of any kind, whether written, oral, implied or statutory, including, without limitation, the warranties of merchantability, fitness for a particular purpose and non-infringement.  Intel makes no claims as to the accuracy, completeness, timeliness or fitness for any particular purpose of any information and data contained herein or of any results produced by this document, or that such information, data or results will be error-free.  You are solely responsible for verifying that such information, data and results are accurate.  If you use the results generated by this document for any purpose, you do so at your own discretion and risk and Intel will not be liable for any loss, damage or inconvenience caused as a result of your use of this document. Intel reserves the right, at its sole discretion, to modify or change any part of this document at any time.  Intel and the Intel logo are trademarks of Intel Corporation in the U.S. and/or other countries. Intel Corporation"

    worksheet.column_dimensions['A'].width = 160
    worksheet['A1'].font = Font(size=24, bold=True)
    worksheet['A1'].value = 'Legal'
    worksheet['A2'].alignment = Alignment(wrapText=True)
    worksheet['A2'].font = Font(size=20)
    worksheet['A2'].value = text

    xfile.save(Reports_path)


def gen_mem_text(title):
    global last_row, table_start, table_end
    Get_input_row()

    text = '''1. Use identical DIMM types throughout the platform:
    - Same size, speed, and number of ranks
2. Maximize the same number of channels populated in each memory controller
3. Use a "balanced" platform configuration:
    - All available memory channels populated equally
    - Identical DIMMs in all locations (size/speed/rank)
4. Use a "near-balanced" platform configuration:
    - All available memory channels and sockets populated equally
    - Identical DIMMs in each "row", but different sized DIMMs in row #1 vs. row #2'''

    worksheet['A' + str(last_row)].font = Font(size=14, bold=True)
    worksheet['A' + str(last_row)] = title
    last_row = last_row + 1
    table_start = last_row
    worksheet['A' + str(last_row)] = 'Guidelines for optimizing Memory Performance:'
    last_row = last_row + 1
    worksheet['A' + str(last_row)].alignment = Alignment(wrapText=True)
    worksheet['A' + str(last_row)] = text

    row_input = last_row


def parsing():
    global Reports_path

    parser = argparse.ArgumentParser(description='Generate svr_info html report.')
    parser.add_argument('-i', '--input', help='Raw svr_info log', required=True)
    parser.add_argument('-c', '--customer', help='Raw svr_info log', required=True)
    parser.add_argument('-o', '--output', help='Enter the output excel report path')
    args = vars(parser.parse_args())

    s = Svrinfo(args['input'])
    customer = Svrinfo(args['customer'])

    if not args['output'] == None:
        output_path = args['output']
        Reports_path = output_path + Reports_path

    return s, customer


def main_binary(input_arg, customer, output_path=False):
    global Reports_path

    s = Svrinfo(input_arg)
    customer = Svrinfo(customer)

    if not output_path == False:
        Reports_path = output_path + Reports_path

    main(s, customer)


def main(s, customer):
    adding_sheet('Legal')
    adding_sheet('ServerInfo')

    gen_legal_text()
    # load excel
    global xfile, worksheet
    xfile = openpyxl.load_workbook(Reports_path)
    worksheet = xfile['ServerInfo']

    global table_count
    table_count = 1

    with open("svr_info.yml", 'r') as ymlfile:
        cfg = yaml.load(ymlfile, Loader=yaml.SafeLoader)

    for info in cfg['svr_info_list']:
        try:
            if info == 'Host Name and Time':
                gen_report('Host Name and Time', 'Host Info', s.get_sys(), customer.get_sys())
            elif info == 'System Details':
                gen_report('System Details', 'System Info', s.get_sysd(), customer.get_sysd())
            elif info == 'CPU Details':
                gen_report('CPU Details', 'CPU Info', s.get_cpu(), customer.get_cpu())
            elif info == 'Memory Details':
                gen_report('Memory Details', 'Memory Info', s.get_mem(), customer.get_mem())
            elif info == 'Memory Topology for hosts: dut.':
                gen_report_dimm('Memory Topology for hosts: dut.', s.get_dimms(), customer.get_dimms())
            elif info == 'Memory Bandwidth -vs- Latency Performance Chart :':
                gen_mem_text('Memory Performance')
                gen_report_graph('Memory Bandwidth -vs- Latency Performance Chart :', s.get_loadlat(), customer.get_loadlat())
            elif info == 'Network Details':
                gen_report_net('Network Details', 'Network Info', s.get_net(), customer.get_net())
            elif info == 'BIOS Details':
                gen_report('BIOS Details', 'BIOS Info', s.get_bios_info_ek(), customer.get_bios_info_ek())
            elif info == 'CMDLINE Details':
                gen_report_single_result('CMDLINE Details', 'CMDLINE Info', s.get_cmdline(), customer.get_cmdline())
            elif info == 'GCC Version':
                gen_report_single_result('GCC Version', 'GCC Info', s.get_gccver(), customer.get_gccver())
            elif info == 'MSR':
                # gen_report_single_result('MSR', 'MSR Info', s.get_MSR(), customer.get_MSR())
                gen_report('MSR', 'MSR Info', s.get_msr(), customer.get_msr())
            elif info == 'JAVA Version':
                gen_report_single_result('JAVA Version', 'JAVA Info', s.get_javaver(), customer.get_javaver())
            elif info == 'PERF PLIMIT':
                gen_report('PERF PLIMIT', 'PERF PLIMIT Info', s.get_PERF_PLIMIT(), customer.get_PERF_PLIMIT())
            elif info == 'CPU IDLE':
                gen_report_single_result('CPU IDLE', 'CPU IDLE Info', s.get_CPU_IDLE(), customer.get_CPU_IDLE())
        except:
            print('not match')
    # before closing file, make the default page to Readme
    xfile.active = 0

    # save file
    xfile.save(Reports_path)


if __name__ == "__main__":
    s, customer = parsing()
    main(s, customer)